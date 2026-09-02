"""Write a set of tables out as CSV, Excel or DuckDB.

A property is never one table - it is a property, its owners, its charges, the
people named on them, its sale history and the building it sits in - so every
exporter here takes the whole set at once, keyed by table name, and each format
does the obvious thing with it:

    CSV      one file per table, in a folder named after the address
    Excel    one sheet per table, in a single workbook
    DuckDB   one table each, which is the shape they were already in

Column order follows the schema in yaybo.store where there is one, so an export
of the same table twice running has the same columns in the same places.
Anything the schema does not name - the widened ejer_1_navn columns, a column
somebody's SQL invented - follows after it rather than being dropped.
"""

from __future__ import annotations

import csv
import json
import re
from datetime import date, datetime
from pathlib import Path

from yaybo import store
from yaybo.register.address import slugify

EXPORT_DIR = Path("exports")


def stamp() -> str:
    return datetime.now().strftime("%Y%m%dT%H%M%S")


def _dest(outdir: Path | None) -> Path:
    destination = Path(outdir) if outdir else EXPORT_DIR
    destination.mkdir(parents=True, exist_ok=True)
    return destination


def columns(name: str, rows: list[dict]) -> list[str]:
    """The column order for one table: the schema's first, then anything else."""
    known = [column for column, _ in store.TABLES.get(name, {}).get("columns", [])]
    seen: dict[str, None] = {}
    for row in rows:
        for key in row:
            seen.setdefault(key, None)
    return [c for c in known if c in seen] + [c for c in seen if c not in known]


def _cell(value):
    """Flatten one value into something a spreadsheet cell can hold."""
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime):
            return value.isoformat(sep=" ")
        return value.isoformat()
    return value


def _filled(tables: dict[str, list[dict]]) -> dict[str, list[dict]]:
    """Drop the tables with nothing in them - an empty sheet is just noise."""
    return {name: rows for name, rows in tables.items() if rows}


# ── CSV ─────────────────────────────────────────────────────────────────


def export_csv(
    tables: dict[str, list[dict]], stem: str, *, outdir: Path | None = None
) -> list[Path]:
    """One file per table. A single table writes one file, not a folder."""
    tables = _filled(tables)
    if not tables:
        return []
    destination = _dest(outdir)
    slug = slugify(stem) or "yaybo"
    at = stamp()

    if len(tables) == 1:
        name, rows = next(iter(tables.items()))
        return [_write_csv(destination / f"{slug}-{name}-{at}.csv", name, rows)]

    folder = destination / f"{slug}-{at}"
    folder.mkdir(parents=True, exist_ok=True)
    return [
        _write_csv(folder / f"{name}.csv", name, rows) for name, rows in tables.items()
    ]


def _write_csv(path: Path, name: str, rows: list[dict]) -> Path:
    fields = columns(name, rows)
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(fields)
        writer.writerows([[_cell(row.get(f)) for f in fields] for row in rows])
    return path


# ── Excel ───────────────────────────────────────────────────────────────

# Excel's own rules for a sheet name, which openpyxl will not enforce for you.
SHEET_ILLEGAL = re.compile(r"[\[\]:*?/\\]")


def export_xlsx(
    tables: dict[str, list[dict]], stem: str, *, outdir: Path | None = None
) -> Path | None:
    """One workbook, one sheet per table, with the header row frozen."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    tables = _filled(tables)
    if not tables:
        return None
    path = _dest(outdir) / f"{slugify(stem) or 'yaybo'}-{stamp()}.xlsx"

    book = Workbook()
    book.remove(book.active)  # a fresh workbook opens with one sheet we do not want
    for name, rows in tables.items():
        sheet = book.create_sheet(SHEET_ILLEGAL.sub("_", name)[:31])
        fields = columns(name, rows)
        sheet.append(fields)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in rows:
            sheet.append([_cell(row.get(field)) for field in fields])

        # Enough width to read the column, capped so one long free-text cell
        # cannot push the rest of the sheet off the screen.
        for index, field in enumerate(fields, start=1):
            widest = max(
                [len(str(field))]
                + [len(str(_cell(row.get(field)))) for row in rows[:200]]
            )
            sheet.column_dimensions[get_column_letter(index)].width = min(widest + 2, 60)
        sheet.freeze_panes = "A2"

    book.save(path)
    return path


# ── DuckDB ──────────────────────────────────────────────────────────────


def export_duckdb(
    tables: dict[str, list[dict]], stem: str, *, outdir: Path | None = None
) -> Path | None:
    """A standalone database of just these rows, typed from the schema.

    Not the same thing as the database a run accumulates into: this one is a
    snapshot to hand to somebody, so it is written fresh each time rather than
    replacing rows in place.
    """
    import duckdb

    tables = _filled(tables)
    if not tables:
        return None
    path = _dest(outdir) / f"{slugify(stem) or 'yaybo'}-{stamp()}.duckdb"

    with duckdb.connect(str(path)) as db:
        for name, rows in tables.items():
            fields = columns(name, rows)
            declared = dict(store.TABLES.get(name, {}).get("columns", []))
            typed = [
                (field, declared.get(field) or _guess(rows, field)) for field in fields
            ]
            definition = ", ".join(f'"{field}" {sort}' for field, sort in typed)
            db.execute(f'CREATE TABLE "{name}" ({definition})')
            holes = ", ".join("?" * len(typed))
            db.executemany(
                f'INSERT INTO "{name}" VALUES ({holes})',
                [
                    [
                        # A column the schema names arrived as scraped text and
                        # has to be talked into its type; one it does not name
                        # came out of a query already typed, and re-reading it
                        # through the register's Danish number rules would only
                        # break it.
                        store.coerce(row.get(field), sort)
                        if field in declared
                        else _plain(row.get(field))
                        for field, sort in typed
                    ]
                    for row in rows
                ],
            )
    return path


def _plain(value):
    """A value from a query, as DuckDB will take it back."""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _guess(rows: list[dict], field: str) -> str:
    """A column type for something the schema never named."""
    for row in rows:
        value = row.get(field)
        if value is None or value == "":
            continue
        if isinstance(value, bool):
            return store.BOOLEAN
        if isinstance(value, int):
            return store.INTEGER
        if isinstance(value, float):
            return store.DECIMAL
        if isinstance(value, datetime):
            return "TIMESTAMP"
        if isinstance(value, date):
            return store.DATE
        break
    return store.TEXT


FORMATS = {
    "CSV": export_csv,
    "Excel (XLSX)": export_xlsx,
    "DuckDB": export_duckdb,
}
